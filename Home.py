# -*- coding: utf-8 -*-
"""
Created on Tue Feb 28 21:35:46 2023

@author: YOGB
"""
import streamlit as st
import openpyxl
import pandas as pd

st.header("Coba")

if st.button("Click"):
    wb_out = openpyxl.Workbook()
    wb_out.remove(worksheet=wb_out['Sheet'])

    sheet_bhmgr = wb_out.create_sheet('borehole_manager')
    for i in range(5):
        sheet_bhmgr.cell(i+1,1).value = i
    
    wb_out.save(filename="contoh.xlsx")
    
    dt = pd.read_excel("contoh.xlsx", sheet_name="borehole_manager")

    st.write(dt)
